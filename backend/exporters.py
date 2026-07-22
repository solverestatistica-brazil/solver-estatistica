"""Exportadores Solver — PDF acadêmico claro, Excel e gráficos.

v3: PDF acadêmico claro, em A4 retrato, com capa e margens baseadas na ABNT
NBR 14724. Excel mantém cabeçalho escuro + corpo claro para impressão.
"""

from __future__ import annotations

import base64
import io
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from zoneinfo import ZoneInfo
from xml.sax.saxutils import escape

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    BaseDocTemplate, Frame, Image, KeepTogether, NextPageTemplate, PageBreak, PageTemplate,
    Paragraph, Spacer, Table, TableStyle,
)

from statistics_engine import analyze, _register_chart_fonts, CHART_TITLE_FONT
from provenance import build_provenance

FONT_DIR = Path(__file__).resolve().parent / "fonts"
FONT_TEXT = "OpenSans"
FONT_TEXT_SEMIBOLD = "OpenSans-Semibold"
FONT_TEXT_BOLD = "OpenSans-Bold"
FONT_HEADING = "Exo2"
FONT_HEADING_SEMIBOLD = "Exo2-Semibold"
FONT_HEADING_BOLD = "Exo2-Bold"
FONT_HEADING_BLACK = "Exo2-ExtraBold"

for font_name, file_name in (
    (FONT_TEXT, "OpenSans-Regular.ttf"),
    (FONT_TEXT_SEMIBOLD, "OpenSans-Semibold.ttf"),
    (FONT_TEXT_BOLD, "OpenSans-Bold.ttf"),
    (FONT_HEADING, "Exo2-Regular.ttf"),
    (FONT_HEADING_SEMIBOLD, "Exo2-Semibold.ttf"),
    (FONT_HEADING_BOLD, "Exo2-Bold.ttf"),
    (FONT_HEADING_BLACK, "Exo2-ExtraBold.ttf"),
):
    if font_name not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont(font_name, str(FONT_DIR / file_name)))

pdfmetrics.registerFontFamily(
    FONT_TEXT,
    normal=FONT_TEXT,
    bold=FONT_TEXT_BOLD,
    italic=FONT_TEXT,
    boldItalic=FONT_TEXT_BOLD,
)
pdfmetrics.registerFontFamily(
    FONT_HEADING,
    normal=FONT_HEADING,
    bold=FONT_HEADING_BOLD,
    italic=FONT_HEADING,
    boldItalic=FONT_HEADING_BOLD,
)

STATUS_LABELS = {
    "ok": "OK",
    "violado": "VIOLADO",
    "atencao": "ATENÇÃO",
    "indeterminado": "INDETERMINADO",
}


def _status_label(value: Any) -> str:
    """Converte o status técnico interno em um rótulo legível no documento."""
    raw = str(value or "").strip()
    return STATUS_LABELS.get(raw.lower(), raw.upper() or "—")

# ============================================================================
# Paleta v3 — documento claro e adequado para leitura/impressão
# ============================================================================
CANVAS = colors.HexColor("#FFFFFF")
CANVAS_ELEVATED = colors.HexColor("#F8FAF9")
CANVAS_ELEVATED_2 = colors.HexColor("#E8F2EC")
BORDER = colors.HexColor("#CFDDD4")
BORDER_BRAND = colors.HexColor("#1B6E3D")

TEXT_D1 = colors.HexColor("#17221B")
TEXT_D2 = colors.HexColor("#4C5C52")
TEXT_D3 = colors.HexColor("#6B756F")

BRAND = colors.HexColor("#187A43")
BRAND_HI = colors.HexColor("#2E9B5F")
BRAND_DEEP = colors.HexColor("#0F5132")
BRAND_DIM = colors.HexColor("#E8F5ED")

SUCCESS = BRAND
SUCCESS_DIM = colors.HexColor("#E7F5EC")
WARNING = colors.HexColor("#A35F00")
WARNING_DIM = colors.HexColor("#FFF3D9")
ERROR = colors.HexColor("#B42318")
ERROR_DIM = colors.HexColor("#FDECEA")
NEUTRAL = colors.HexColor("#66706A")
NEUTRAL_DIM = colors.HexColor("#F1F3F2")

# Excel — cabeçalho escuro, corpo claro para impressão.
HEX = {
    "brand_deep": "166534",
    "brand": "22C55E",
    "canvas": "0A0A0A",
    "surface": "FFFFFF",
    "surface_line": "E5E7EB",
    "surface_subtle": "F4F6F5",
    "success": "16A34A",
    "success_tint": "DCFCE7",
    "warning": "D97706",
    "warning_tint": "FEF3C7",
    "neutral": "737373",
    "neutral_tint": "F5F5F5",
    "text_l1": "0F1F14",
    "text_l2": "525252",
    "white": "FFFFFF",
}

DESIGN_LABELS = {
    "DIC": "Inteiramente Casualizado (DIC)",
    "DBC": "Blocos Casualizados (DBC)",
    "DQL": "Quadrado Latino (DQL)",
}
BRASILIA_TZ = ZoneInfo("America/Sao_Paulo")

def _brasilia_now() -> datetime:
    return datetime.now(BRASILIA_TZ)

def _format_brasilia_timestamp(value: Any) -> str:
    """Normaliza timestamps de proveniência para a apresentação local."""
    if not value:
        return "—"
    try:
        instant = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        if instant.tzinfo is None:
            instant = instant.replace(tzinfo=timezone.utc)
        return instant.astimezone(BRASILIA_TZ).strftime("%d/%m/%Y às %H:%M BRT")
    except (TypeError, ValueError):
        return str(value)

TYPE_LABELS = {
    "single": "fator único",
    "factorial": "fatorial",
    "split_plot": "parcelas subdivididas",
    "regression": "regressão direta",
}


def _fmt(value: Any) -> str:
    if value is None:
        return "—"
    if isinstance(value, float):
        return f"{value:.4f}".replace(".", ",")
    return str(value)


def _fmt_pct(value: Any) -> str:
    if value is None:
        return "—"
    return f"{value:.2f}%".replace(".", ",")


def _draw_logo(canvas_obj, logo_x: float, logo_y: float, logo_size: float) -> None:
    """Desenha a marca vetorial sem depender de imagem externa."""
    view = 40.0
    scale = logo_size / view

    def _sp(x: float, y: float) -> Tuple[float, float]:
        return logo_x + x * scale, logo_y + logo_size - y * scale

    canvas_obj.setStrokeColor(BRAND)
    canvas_obj.setLineWidth(1.6 * scale)
    canvas_obj.setFillColor(colors.white)
    canvas_obj.roundRect(logo_x, logo_y, logo_size, logo_size, 10 * scale, fill=1, stroke=1)

    canvas_obj.setStrokeColor(BRAND_HI)
    canvas_obj.setLineWidth(2.4 * scale)
    canvas_obj.setLineCap(1)
    canvas_obj.setLineJoin(1)

    trend = canvas_obj.beginPath()
    for i, (px, py) in enumerate([(10, 25), (16, 18), (20, 21), (28, 12)]):
        cx, cy = _sp(px, py)
        (trend.moveTo if i == 0 else trend.lineTo)(cx, cy)
    canvas_obj.drawPath(trend, stroke=1, fill=0)

    arrow = canvas_obj.beginPath()
    for i, (px, py) in enumerate([(23, 12), (28, 12), (28, 17)]):
        cx, cy = _sp(px, py)
        (arrow.moveTo if i == 0 else arrow.lineTo)(cx, cy)
    canvas_obj.drawPath(arrow, stroke=1, fill=0)


COVER_BG = colors.HexColor("#FAFBF9")
COVER_SIDEBAR = colors.HexColor("#0F4A2E")
COVER_ACCENT = colors.HexColor("#55B06B")
COVER_WATERMARK = colors.HexColor("#A6C3B0")  # verde acinzentado suave p/ marca d'água
COVER_TITLE_DARK = colors.HexColor("#111111")
COVER_TITLE_GREEN = colors.HexColor("#2EA052")
COVER_SUB1 = colors.HexColor("#3A4A3D")
COVER_SUB2 = colors.HexColor("#5A6B5D")
COVER_BADGE_BG = colors.HexColor("#E6F2EA")
COVER_BADGE_TEXT = colors.HexColor("#2C3D30")


def _cover_icon_bars(c, cx: float, cy: float) -> None:
    # Traço 1.3 e caixa óptica ~0.22 cm, padronizados com os demais selos.
    c.setStrokeColor(BRAND_DEEP)
    c.setLineWidth(1.3)
    c.setLineCap(1)
    c.setLineJoin(1)
    c.line(cx - 0.17 * cm, cy - 0.17 * cm, cx - 0.17 * cm, cy + 0.02 * cm)
    c.line(cx,             cy - 0.17 * cm, cx,             cy + 0.11 * cm)
    c.line(cx + 0.17 * cm, cy - 0.17 * cm, cx + 0.17 * cm, cy + 0.20 * cm)


def _cover_icon_target(c, cx: float, cy: float) -> None:
    # Traço 1.3 e caixa óptica ~0.22 cm, padronizados com os demais selos.
    c.setStrokeColor(BRAND_DEEP)
    c.setLineWidth(1.3)
    c.setLineCap(1)
    c.circle(cx, cy, 0.185 * cm, stroke=1, fill=0)
    c.circle(cx, cy, 0.065 * cm, stroke=1, fill=0)
    for dx, dy in ((0, 1), (0, -1), (1, 0), (-1, 0)):
        c.line(cx + dx * 0.185 * cm, cy + dy * 0.185 * cm,
               cx + dx * 0.24 * cm,  cy + dy * 0.24 * cm)


def _cover_icon_bulb(c, cx: float, cy: float) -> None:
    """Lâmpada — símbolo de ideia/decisão inteligente (combina com "Decisões inteligentes")."""
    # Traço 1.3 e caixa óptica ~0.22 cm, padronizados com os demais selos.
    c.setStrokeColor(BRAND_DEEP)
    c.setLineWidth(1.3)
    c.setLineCap(1)
    c.setLineJoin(1)
    # Bulbo
    bulb_cy = cy + 0.055 * cm
    r = 0.155 * cm
    c.circle(cx, bulb_cy, r, stroke=1, fill=0)
    # Gargalo (ombros do bulbo até a base)
    neck = 0.085 * cm
    base_top = bulb_cy - r + 0.005 * cm
    c.line(cx - neck, base_top, cx - neck, base_top - 0.045 * cm)
    c.line(cx + neck, base_top, cx + neck, base_top - 0.045 * cm)
    # Rosca da base (duas ranhuras)
    by = base_top - 0.045 * cm
    c.line(cx - neck, by, cx + neck, by)
    c.line(cx - neck * 0.7, by - 0.06 * cm, cx + neck * 0.7, by - 0.06 * cm)
    # Filamento (ziguezague) dentro do bulbo
    fil = c.beginPath()
    fil.moveTo(cx - 0.07 * cm, bulb_cy + 0.045 * cm)
    fil.lineTo(cx - 0.025 * cm, bulb_cy - 0.028 * cm)
    fil.lineTo(cx + 0.018 * cm, bulb_cy + 0.035 * cm)
    fil.lineTo(cx + 0.063 * cm, bulb_cy - 0.028 * cm)
    c.drawPath(fil, stroke=1, fill=0)


def _draw_cover_page(canvas_obj, doc) -> None:
    """Capa institucional: título grande à esquerda, sidebar com ano, selos no rodapé."""
    canvas_obj.saveState()
    width, height = A4

    # Fundo geral
    canvas_obj.setFillColor(COVER_BG)
    canvas_obj.rect(0, 0, width, height, fill=1, stroke=0)

    # Faixa lateral verde escura
    sidebar_w = 2.2 * cm
    canvas_obj.setFillColor(COVER_SIDEBAR)
    canvas_obj.rect(0, 0, sidebar_w, height, fill=1, stroke=0)

    # Barra clara acima do ano
    canvas_obj.setFillColor(COVER_ACCENT)
    canvas_obj.rect(sidebar_w / 2 - 0.12 * cm, 5.5 * cm, 0.24 * cm, 2.2 * cm, fill=1, stroke=0)

    # Ano (dinâmico) — 20 sobre 24
    canvas_obj.setFillColor(colors.white)
    canvas_obj.setFont(FONT_TEXT, 14)
    year_now = str(getattr(doc, "cover_year", _brasilia_now().year))
    canvas_obj.drawCentredString(sidebar_w / 2, 4.7 * cm, year_now[:2])
    canvas_obj.drawCentredString(sidebar_w / 2, 4.0 * cm, year_now[2:])

    # Grafismo decorativo como marca d'água suave. Envolto em save/restore para que a
    # transparência não afete o logo, o título e os selos desenhados depois.
    canvas_obj.saveState()

    # Pontos decorativos no canto superior direito (bem sutis)
    canvas_obj.setFillAlpha(0.30)
    canvas_obj.setFillColor(COVER_WATERMARK)
    for i in range(10):
        for j in range(4):
            canvas_obj.circle(
                width - 2.5 * cm - i * 0.26 * cm,
                height - 2.5 * cm - j * 0.26 * cm,
                0.7, fill=1, stroke=0,
            )

    # Círculos concêntricos à direita (traço fino e translúcido)
    cx, cy = width - 5.5 * cm, height / 2 + 0.5 * cm
    canvas_obj.setStrokeAlpha(0.45)
    canvas_obj.setStrokeColor(COVER_WATERMARK)
    canvas_obj.setLineWidth(0.5)
    for r in (5.5 * cm, 4.2 * cm, 3.0 * cm):
        canvas_obj.circle(cx, cy, r, stroke=1, fill=0)
    canvas_obj.setFillAlpha(0.5)
    canvas_obj.setFillColor(COVER_WATERMARK)
    canvas_obj.circle(cx, cy, 1.6, fill=1, stroke=0)

    # Curva decorativa passando pelo centro dos círculos (bem suave)
    p = canvas_obj.beginPath()
    p.moveTo(sidebar_w + 1.5 * cm, 9.0 * cm)
    p.curveTo(sidebar_w + 6.0 * cm, 16.0 * cm, cx - 4.0 * cm, cy - 2.0 * cm, cx, cy)
    canvas_obj.setStrokeAlpha(0.4)
    canvas_obj.setStrokeColor(COVER_WATERMARK)
    canvas_obj.setLineWidth(0.7)
    canvas_obj.drawPath(p, stroke=1, fill=0)

    canvas_obj.restoreState()

    # ------------------------------------------------------------------ LOGO
    logo_x, logo_y = sidebar_w + 1.8 * cm, height - 4.2 * cm
    _draw_logo(canvas_obj, logo_x, logo_y, 1.1 * cm)
    canvas_obj.setFillColor(BRAND_DEEP)
    canvas_obj.setFont(FONT_HEADING_BOLD, 17)
    canvas_obj.drawString(logo_x + 1.35 * cm, logo_y + 0.48 * cm, "SOLVER")
    canvas_obj.setFillColor(TEXT_D3)
    canvas_obj.setFont(FONT_TEXT_SEMIBOLD, 6.4)
    canvas_obj.drawString(logo_x + 1.35 * cm, logo_y + 0.15 * cm, "INTELLIGENCE FOR FIELD TRIALS")

    # ------------------------------------------------------------- TÍTULO
    title_x = sidebar_w + 1.8 * cm
    canvas_obj.setFillColor(COVER_TITLE_DARK)
    canvas_obj.setFont(FONT_HEADING_BLACK, 45)
    canvas_obj.drawString(title_x, height - 10.8 * cm, "RELATÓRIO")
    canvas_obj.setFillColor(COVER_TITLE_GREEN)
    canvas_obj.drawString(title_x, height - 12.3 * cm, "ESTATÍSTICO")

    # Divisor
    canvas_obj.setFillColor(COVER_TITLE_GREEN)
    canvas_obj.rect(title_x, height - 13.1 * cm, 4.6 * cm, 0.14 * cm, fill=1, stroke=0)

    # Subtítulos dinâmicos
    subtitle_main = getattr(doc, "cover_design_label", "") or "Análise estatística experimental"
    subtitle_second = getattr(doc, "cover_type_label", "") or ""
    canvas_obj.setFillColor(COVER_SUB1)
    canvas_obj.setFont(FONT_TEXT, 12)
    canvas_obj.drawString(title_x, height - 13.9 * cm, subtitle_main)
    if subtitle_second:
        canvas_obj.setFillColor(COVER_SUB2)
        canvas_obj.drawString(title_x, height - 14.55 * cm, subtitle_second)

    # ------------------------------------------------------- RODAPÉ / SELOS
    footer_h = 2.4 * cm
    canvas_obj.setFillColor(colors.white)
    canvas_obj.rect(sidebar_w, 0, width - sidebar_w, footer_h, fill=1, stroke=0)
    canvas_obj.setStrokeColor(BORDER)
    canvas_obj.setLineWidth(0.6)
    canvas_obj.line(sidebar_w, footer_h, width, footer_h)

    badges = (
        ("DADOS", "CONFIÁVEIS", _cover_icon_bars),
        ("ANÁLISES", "PRECISAS", _cover_icon_target),
        ("DECISÕES", "INTELIGENTES", _cover_icon_bulb),
    )
    zone_w = (width - sidebar_w) / len(badges)
    for i, (t1, t2, icon_fn) in enumerate(badges):
        zone_cx = sidebar_w + zone_w * (i + 0.5)
        icon_cx = zone_cx - 1.4 * cm
        icon_cy = footer_h / 2
        canvas_obj.setFillColor(COVER_BADGE_BG)
        canvas_obj.circle(icon_cx, icon_cy, 0.45 * cm, fill=1, stroke=0)
        icon_fn(canvas_obj, icon_cx, icon_cy)
        canvas_obj.setFillColor(COVER_BADGE_TEXT)
        canvas_obj.setFont(FONT_TEXT_BOLD, 7.5)
        canvas_obj.drawString(icon_cx + 0.65 * cm, icon_cy + 0.10 * cm, t1)
        canvas_obj.drawString(icon_cx + 0.65 * cm, icon_cy - 0.35 * cm, t2)

    canvas_obj.restoreState()

def _draw_content_header_footer(canvas_obj, doc) -> None:
    """Cabeçalho e rodapé das páginas textuais em A4 retrato."""
    canvas_obj.saveState()
    width, height = A4
    canvas_obj.setFillColor(CANVAS)
    canvas_obj.rect(0, 0, width, height, fill=1, stroke=0)

    _draw_logo(canvas_obj, 3.0 * cm, height - 1.75 * cm, 0.75 * cm)
    canvas_obj.setFillColor(TEXT_D1)
    canvas_obj.setFont(FONT_HEADING_BOLD, 9.5)
    canvas_obj.drawString(3.95 * cm, height - 1.28 * cm, "SOLVER ESTATÍSTICA")
    canvas_obj.setFont(FONT_TEXT, 7.2)
    canvas_obj.setFillColor(TEXT_D3)
    canvas_obj.drawString(3.95 * cm, height - 1.58 * cm, "Relatório estatístico experimental")
    canvas_obj.setFont(FONT_TEXT, 7.2)
    canvas_obj.setFillColor(BRAND)
    canvas_obj.drawRightString(
        width - 2.0 * cm,
        height - 1.42 * cm,
        getattr(doc, "generated_at_brasilia_display", _format_brasilia_timestamp(_brasilia_now())),
    )
    canvas_obj.setStrokeColor(BORDER)
    canvas_obj.setLineWidth(0.6)
    canvas_obj.line(3.0 * cm, height - 2.05 * cm, width - 2.0 * cm, height - 2.05 * cm)
    canvas_obj.line(3.0 * cm, 1.45 * cm, width - 2.0 * cm, 1.45 * cm)
    canvas_obj.setFont(FONT_TEXT, 7.2)
    canvas_obj.setFillColor(TEXT_D3)
    canvas_obj.drawString(
        3.0 * cm,
        1.0 * cm,
        f"Solver {getattr(doc, 'engine_version', '—')} · commit {getattr(doc, 'commit_short', '—')} · revisão técnica obrigatória.",
    )
    canvas_obj.drawRightString(width - 2.0 * cm, 1.0 * cm, f"Página {max(1, doc.page - 1)}")
    canvas_obj.restoreState()


def _kpi_card(label: str, value: str, sub: str) -> Table:
    """Card claro e compacto para a largura útil do A4 retrato."""
    label_style = ParagraphStyle(
        "KpiLabel", fontName=FONT_HEADING_BOLD, fontSize=6.8,
        textColor=TEXT_D3, leading=9,
    )
    value_style = ParagraphStyle(
        "KpiValue", fontName=FONT_HEADING_BOLD, fontSize=15,
        textColor=TEXT_D1, leading=17, spaceBefore=4,
    )
    sub_style = ParagraphStyle(
        "KpiSub", fontName=FONT_HEADING_BOLD, fontSize=7,
        textColor=BRAND, spaceBefore=3,
    )
    card = Table(
        [[Paragraph(label.upper(), label_style)],
         [Paragraph(value, value_style)],
         [Paragraph(sub, sub_style)]],
        colWidths=[5.0 * cm],
    )
    card.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.8, BORDER),
        ("ROUNDEDCORNERS", [10, 10, 10, 10]),
        ("BACKGROUND", (0, 0), (-1, -1), CANVAS_ELEVATED),
        ("TOPPADDING", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
        ("LEFTPADDING", (0, 0), (-1, -1), 9),
        ("RIGHTPADDING", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    return card


def _kpi_cards(result: Dict[str, Any]) -> Table:
    cv = result.get("anova", {}).get("cv")
    cv_label = result.get("anova", {}).get("cv_label", "Indisponível")
    n_rows = result.get("meta", {}).get("n_rows")
    best = (result.get("means") or {}).get("best")
    best_label = best.get("treatment") if best else "—"
    best_mean = f"Média {_fmt(best.get('mean'))}" if best else "—"

    card1 = _kpi_card("CV experimental", _fmt_pct(cv) if cv is not None else "—", cv_label)
    card2 = _kpi_card("Linhas analisadas", str(n_rows if n_rows is not None else "—"), "Observações")
    card3 = _kpi_card("Melhor tratamento", str(best_label), best_mean)

    card_w = 5.0 * cm
    gap_w = 0.5 * cm
    layout = Table([[card1, "", card2, "", card3]], colWidths=[card_w, gap_w, card_w, gap_w, card_w])
    layout.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    return layout


def _sig_colors(value: Optional[str]):
    """Fundo + cor do texto para o badge de significância."""
    if value == "1%":
        return SUCCESS_DIM, BRAND
    if value == "5%":
        return WARNING_DIM, WARNING
    if value in (None, "—", "-", "ns"):
        return NEUTRAL_DIM, NEUTRAL
    return None, None


def _styled_table(
    rows: List[List[Any]], sig_col: Optional[int] = None, col_widths: Optional[List[float]] = None,
    right_align_from: Optional[int] = 1, row_padding: float = 8,
) -> Table:
    """Tabela clara, repetível entre páginas e compatível com impressão A4."""
    table = Table(rows, repeatRows=1, colWidths=col_widths)
    n_rows = len(rows)
    style = [
        # header
        ("BACKGROUND", (0, 0), (-1, 0), CANVAS_ELEVATED_2),
        ("TEXTCOLOR", (0, 0), (-1, 0), BRAND_DEEP),
        ("FONTNAME", (0, 0), (-1, 0), FONT_HEADING_BOLD),
        ("FONTSIZE", (0, 0), (-1, -1), 7.2),
        # corpo
        ("TEXTCOLOR", (0, 1), (-1, -1), TEXT_D1),
        ("BACKGROUND", (0, 1), (-1, -1), CANVAS_ELEVATED),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, CANVAS_ELEVATED]),
        # linhas separadoras
        ("LINEBELOW", (0, 0), (-1, n_rows - 2), 0.5, BORDER),
        ("LINEBELOW", (0, 0), (-1, 0), 1.4, BRAND),
        ("BOX", (0, 0), (-1, -1), 0.8, BORDER),
        ("ROUNDEDCORNERS", [8, 8, 8, 8]),
        # alinhamento e padding
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), min(row_padding, 5.5)),
        ("BOTTOMPADDING", (0, 0), (-1, -1), min(row_padding, 5.5)),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("FONTNAME", (0, 1), (-1, -1), FONT_TEXT),
        ("FONTNAME", (0, 1), (0, -1), FONT_TEXT_SEMIBOLD),
    ]
    if right_align_from is not None:
        style.append(("ALIGN", (right_align_from, 1), (-1, -1), "RIGHT"))
    if sig_col is not None:
        for row_idx in range(1, n_rows):
            bg, fg = _sig_colors(rows[row_idx][sig_col])
            if bg is None:
                continue
            style.append(("BACKGROUND", (sig_col, row_idx), (sig_col, row_idx), bg))
            style.append(("TEXTCOLOR", (sig_col, row_idx), (sig_col, row_idx), fg))
            style.append(("FONTNAME", (sig_col, row_idx), (sig_col, row_idx), FONT_TEXT_BOLD))
            style.append(("ALIGN", (sig_col, row_idx), (sig_col, row_idx), "CENTER"))
    table.setStyle(TableStyle(style))
    return table


def _intro_text(result: Dict[str, Any]) -> str:
    meta = result.get("meta", {})
    design = meta.get("design")
    analysis_type = meta.get("analysis_type")
    design_label = DESIGN_LABELS.get(design, design or "—")
    type_label = TYPE_LABELS.get(analysis_type, analysis_type or "—")

    if analysis_type == "regression":
        return (
            f"Este relatório apresenta o ajuste de regressão sobre um fator numérico contínuo "
            f"(dose ou concentração), a partir de {meta.get('n_rows')} observações. O objetivo é "
            f"estimar a curva de resposta e o ponto ótimo, e não comparar médias de tratamentos."
        )
    return (
        f"Este relatório apresenta os resultados da análise estatística de um experimento em "
        f"delineamento {design_label}, com estrutura de {type_label}, totalizando {meta.get('n_rows')} "
        f"observações. A análise de variância (ANOVA) avalia, pelo teste F, se há diferença "
        f"estatisticamente significativa entre as fontes de variação testadas, aos níveis de "
        f"5% e 1% de probabilidade."
    )


def _anova_caption() -> str:
    return (
        "Fontes de variação marcadas como <b>1%</b> ou <b>5%</b> apresentam efeito estatisticamente "
        "significativo sobre a variável resposta nesses níveis de probabilidade; <b>ns</b> "
        "(não significativo) indica que não houve evidência estatística de efeito."
    )


def _means_caption(result: Dict[str, Any]) -> Optional[str]:
    comparison = (result.get("means") or {}).get("comparison")
    if not comparison:
        return None
    test_name = comparison.get("test", "TUKEY")
    alpha = comparison.get("alpha", 0.05)
    alpha_pct = f"{alpha * 100:.0f}%".replace(".", ",")
    # [FIX auditoria P1-04] Dunnett so compara cada tratamento contra a testemunha — nunca
    # tratamento contra tratamento. A legenda de "mesma letra" e' invalida aqui (sugeriria
    # relacoes nunca testadas) e a coluna Grupo nem usa mais letras para este teste.
    if test_name == "DUNNETT":
        control = comparison.get("control") or "—"
        legenda = (
            f"Coluna <b>Grupo</b>: 'testemunha' identifica o controle ({control}); 'sig' indica "
            f"diferença estatisticamente significativa contra a testemunha pelo teste de "
            f"<b>Dunnett</b> exato, ao nível de {alpha_pct}; 'ns' indica ausência de diferença "
            f"significativa. O Dunnett <b>não</b> compara tratamentos entre si, apenas cada um "
            f"contra a testemunha."
        )
        note = comparison.get("note") or ""
        if "não informada" in note or "nao informada" in note.lower():
            legenda += f" <b>Atenção:</b> {note}"
        return legenda
    return (
        f"Tratamentos seguidos pela mesma letra na coluna <b>Grupo</b> não diferem estatisticamente "
        f"entre si pelo teste de <b>{test_name.title()}</b>, ao nível de {alpha_pct} de significância."
    )


def build_pdf(payload: Dict[str, Any]) -> bytes:
    """Gera relatório claro em A4 retrato, com capa e margens acadêmicas."""
    result = analyze(payload)
    provenance = result.get("provenance") or build_provenance(payload)
    _NEUTRAL_AUTHOR = "Relatório gerado pelo Solver Estatística"
    _author_display = (str(payload.get("author_name") or "").strip()[:80]) or _NEUTRAL_AUTHOR
    buffer = io.BytesIO()
    page_size = A4
    doc = BaseDocTemplate(
        buffer,
        pagesize=page_size,
        # Margens usuais da ABNT NBR 14724: 3 cm superior/esquerda e 2 cm inferior/direita.
        leftMargin=3.0 * cm,
        rightMargin=2.0 * cm,
        topMargin=3.0 * cm,
        bottomMargin=2.0 * cm,
        title="Relatório Solver Estatística",
        author=_author_display,
        subject="Relatório de análise estatística experimental",
    )
    doc.engine_version = provenance.get("engine_version", "—")
    doc.commit_short = str(provenance.get("git_commit", "—"))[:12]
    doc.generated_at_brasilia_display = _format_brasilia_timestamp(
        provenance.get("generated_at_brasilia") or provenance.get("generated_at_utc")
    )
    _meta_early = result.get("meta") or {}
    _design_early = DESIGN_LABELS.get(
        _meta_early.get("design"), _meta_early.get("design") or "—"
    )
    doc.cover_design_label = f"Delineamento {_design_early}"
    _type_early = TYPE_LABELS.get(
        _meta_early.get("analysis_type"), _meta_early.get("analysis_type") or ""
    )
    doc.cover_type_label = f"Análise de {_type_early}" if _type_early else ""
    doc.cover_year = _brasilia_now().year
    cover_frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id="cover-frame", showBoundary=0)
    content_frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id="content-frame", showBoundary=0)
    doc.addPageTemplates([
        PageTemplate(id="cover", frames=[cover_frame], onPage=_draw_cover_page),
        PageTemplate(id="content", frames=[content_frame], onPage=_draw_content_header_footer),
    ])

    styles = getSampleStyleSheet()
    meta_style = ParagraphStyle(
        "SolverMeta", parent=styles["BodyText"],
        fontName=FONT_TEXT, fontSize=9, leading=13.5, textColor=TEXT_D2, spaceAfter=4,
    )
    h2 = ParagraphStyle(
        "SolverH2", parent=styles["Heading2"],
        fontName=FONT_HEADING_BOLD, fontSize=12, leading=15, textColor=BRAND_DEEP,
        spaceBefore=10, spaceAfter=8, keepWithNext=1,
    )
    body = ParagraphStyle(
        "SolverBody", parent=styles["BodyText"],
        fontName=FONT_TEXT, fontSize=11, leading=16.5, textColor=TEXT_D1,
        alignment=TA_JUSTIFY, firstLineIndent=1.25 * cm,
    )
    body_dim = ParagraphStyle(
        "SolverBodyDim", parent=body, textColor=TEXT_D2, firstLineIndent=0,
    )
    bullet = ParagraphStyle(
        "SolverBullet", parent=body, leftIndent=14, firstLineIndent=0, spaceAfter=4, textColor=TEXT_D2,
    )
    caption = ParagraphStyle(
        "SolverCaption", parent=body_dim,
        fontSize=8.2, leading=11.5, spaceBefore=7, alignment=TA_LEFT,
    )
    cell_small = ParagraphStyle(
        "SolverCellSmall", parent=body_dim,
        fontSize=7.4, leading=9.5, spaceBefore=0, spaceAfter=0, alignment=TA_LEFT,
    )
    cell_ok = ParagraphStyle(
        "SolverCellOk", parent=cell_small, textColor=SUCCESS, fontName=FONT_HEADING_BOLD,
    )
    cell_warn = ParagraphStyle(
        "SolverCellWarn", parent=cell_small, textColor=WARNING, fontName=FONT_HEADING_BOLD,
    )
    cell_neutral = ParagraphStyle(
        "SolverCellNeutral", parent=cell_small, textColor=NEUTRAL, fontName=FONT_HEADING_BOLD,
    )

    def _status_cell(raw_status: Any) -> Paragraph:
        key = str(raw_status or "").strip().lower()
        if key in ("atencao", "atenção", "violado", "warning", "alerta"):
            st = cell_warn
        elif key == "ok":
            st = cell_ok
        else:
            st = cell_neutral
        return Paragraph(escape(_status_label(raw_status)), st)
    tag = ParagraphStyle(
        "SolverTag", parent=body, fontName=FONT_HEADING_BOLD,
        fontSize=8.5, textColor=BRAND, leading=11, firstLineIndent=0, spaceBefore=0, spaceAfter=6,
    )
    # (título e subtítulo da capa são desenhados via canvas em _draw_cover_page)
    meta = result.get("meta", {})
    design_label = DESIGN_LABELS.get(meta.get("design"), meta.get("design") or "—")
    type_label = TYPE_LABELS.get(meta.get("analysis_type"), meta.get("analysis_type") or "—")
    alpha_mode_label = {"auto": "Automático", "fixed": "Fixo"}.get(
        str(meta.get("alpha_mode", "auto")).lower(),
        str(meta.get("alpha_mode", "auto")).title(),
    )
    section_counter = [0]

    def section_heading(title: str) -> Paragraph:
        section_counter[0] += 1
        return Paragraph(f"{section_counter[0]} {escape(title.upper())}", h2)

    story: List[Any] = [
        # Página 1 (capa) — desenhada 100% em canvas via _draw_cover_page.
        # A story só precisa "empurrar" um flowable para o frame existir.
        Spacer(1, 1),
        NextPageTemplate("content"),
        PageBreak(),
        Paragraph(
            f"DELINEAMENTO {escape(str(design_label)).upper()} - "
            f"ANÁLISE DE {escape(str(type_label)).upper()} - "
            f"{meta.get('n_rows')} OBSERVAÇÕES",
            tag,
        ),
        section_heading("Apresentação"),
        Paragraph(_intro_text(result), body),
        Paragraph(
            f"Alfa adotado: <b>{_fmt(meta.get('alpha'))}</b> ({escape(alpha_mode_label)}) - "
            f"Soma de quadrados: <b>tipo {_fmt(meta.get('sum_squares_type'))}</b> - "
            f"Motor: <b>{escape(str(provenance.get('engine_version', '—')))}</b> - "
            f"Commit: <b>{escape(str(provenance.get('git_commit', '—'))[:12])}</b>",
            meta_style,
        ),
        Spacer(1, 0.35 * cm),
        _kpi_cards(result),
        Spacer(1, 0.5 * cm),
        section_heading("Resumo executivo"),
    ]
    for msg in result.get("recommendations", []):
        story.append(Paragraph("• " + msg, bullet))
    story.append(Spacer(1, 0.35 * cm))

    anova_rows = [["FV", "GL", "SQ", "QM", "F calc", "F 5%", "F 1%", "p", "Sig"]]
    for r in result.get("anova", {}).get("table", []):
        anova_rows.append([
            r.get("source"), _fmt(r.get("df")), _fmt(r.get("sum_sq")), _fmt(r.get("mean_sq")),
            _fmt(r.get("f_calc")), _fmt(r.get("f_5")), _fmt(r.get("f_1")), _fmt(r.get("p_value")),
            r.get("significance"),
        ])
    story.append(KeepTogether([
        section_heading("Quadro de ANOVA · Teste F"),
        _styled_table(
            anova_rows,
            sig_col=8,
            col_widths=[3.0 * cm, 0.8 * cm, 1.7 * cm, 1.7 * cm, 1.5 * cm, 1.4 * cm, 1.4 * cm, 1.7 * cm, 0.8 * cm],
        ),
        Paragraph(_anova_caption(), caption),
    ]))
    story.append(Spacer(1, 0.4 * cm))

    means_rows = [["Tratamento", "Média", "n", "DP", "Grupo"]]
    for r in result.get("means", {}).get("treatment_means", []):
        means_rows.append([
            r.get("treatment"), _fmt(r.get("mean")), _fmt(r.get("n")),
            _fmt(r.get("sd")), r.get("group", ""),
        ])
    if len(means_rows) > 1:
        means_block = [
            section_heading("Médias por tratamento"),
            _styled_table(means_rows, col_widths=[4.5 * cm, 3.0 * cm, 1.5 * cm, 3.0 * cm, 4.0 * cm]),
        ]
        m_caption = _means_caption(result)
        if m_caption:
            means_block.append(Paragraph(m_caption, caption))
        story.append(KeepTogether(means_block))
        story.append(Spacer(1, 0.35 * cm))

    means_png = (result.get("means") or {}).get("plot_png_base64")
    if means_png:
        try:
            _img_w = 15.0 * cm
            _img_h = _img_w * 5.2 / 9.0
            _fig = Image(io.BytesIO(base64.b64decode(means_png)), width=_img_w, height=_img_h)
            story.append(KeepTogether([
                section_heading("Gráfico de médias"),
                _fig,
                Paragraph(
                    "Barras de erro = desvio-padrão. Letras iguais indicam tratamentos que não "
                    "diferem pelo teste de médias escolhido.",
                    caption,
                ),
            ]))
            story.append(Spacer(1, 0.35 * cm))
        except Exception:
            # A ausência do gráfico nunca deve invalidar o laudo textual.
            pass

    interaction_png = (result.get("means") or {}).get("interaction_plot_base64")
    if interaction_png:
        try:
            _iw = 15.0 * cm
            _ih = _iw * 5.2 / 9.0
            _ifig = Image(io.BytesIO(base64.b64decode(interaction_png)), width=_iw, height=_ih)
            story.append(KeepTogether([
                section_heading("Gráfico de interação"),
                _ifig,
                Paragraph(
                    "Linhas paralelas sugerem ausência de interação; linhas que divergem ou se "
                    "cruzam indicam interação entre os fatores.",
                    caption,
                ),
            ]))
            story.append(Spacer(1, 0.35 * cm))
        except Exception:
            pass

    comparison = (result.get("means") or {}).get("comparison") or {}
    comparison_rows = comparison.get("comparisons") or []
    if comparison_rows:
        detail_rows: List[List[Any]] = [["Grupo A", "Grupo B", "Diferença", "Dif. crítica", "p ajustado", "Decisão"]]
        for row in comparison_rows:
            detail_rows.append([
                row.get("group_a"), row.get("group_b"), _fmt(row.get("diff")),
                _fmt(row.get("critical_diff")), _fmt(row.get("p_value")),
                "significativo" if row.get("significant") else "não significativo",
            ])
        story.append(KeepTogether([
            section_heading("Comparações detalhadas"),
            _styled_table(
                detail_rows,
                col_widths=[2.3 * cm, 2.3 * cm, 2.5 * cm, 2.5 * cm, 2.3 * cm, 4.1 * cm],
            ),
        ]))
        story.append(Spacer(1, 0.35 * cm))

    factor_comparisons = result.get("factor_comparisons") or []
    if factor_comparisons:
        story.append(section_heading("Comparações por fator"))
        for item in factor_comparisons:
            story.append(Paragraph(
                f"<b>{escape(str(item.get('factor', '—')))}</b> · {escape(str(item.get('test', '—')))} · "
                f"alfa {_fmt(item.get('alpha'))} · erro {escape(str(item.get('error_used', '—')))}",
                caption,
            ))
            rows = [["Nível", "Média", "n", "Grupo"]]
            for level in item.get("levels") or []:
                rows.append([level.get("treatment"), _fmt(level.get("mean")), _fmt(level.get("n")), level.get("group")])
            story.append(_styled_table(rows, col_widths=[5.5 * cm, 3.5 * cm, 2.0 * cm, 5.0 * cm]))
            story.append(Spacer(1, 0.25 * cm))

    breakdown = result.get("interaction_breakdown") or []
    if breakdown:
        story.append(section_heading("Desdobramento da interação"))
        for item in breakdown:
            story.append(Paragraph(
                f"<b>{escape(str(item.get('factor', '—')))} = {escape(str(item.get('level', '—')))}</b> · "
                f"níveis de {escape(str(item.get('sub_factor', '—')))} · {escape(str(item.get('test', '—')))}",
                caption,
            ))
            rows = [["Nível", "Média", "n", "Grupo"]]
            for level in item.get("levels") or []:
                rows.append([level.get("treatment"), _fmt(level.get("mean")), _fmt(level.get("n")), level.get("group")])
            story.append(_styled_table(rows, col_widths=[5.5 * cm, 3.5 * cm, 2.0 * cm, 5.0 * cm]))
            story.append(Spacer(1, 0.25 * cm))

    pressupostos = result.get("pressupostos") or {}
    tests = pressupostos.get("testes") or {}
    if tests:
        story.append(section_heading("Pressupostos da ANOVA"))
        story.append(Paragraph(
            f"Veredito: <b>{escape(_status_label(pressupostos.get('veredito')))}</b> · "
            f"{escape(str(pressupostos.get('resumo', '')))}",
            body,
        ))
        assumption_rows: List[List[Any]] = [["Pressuposto", "Teste", "Status", "Estatística", "p", "Interpretação"]]
        for key, item in tests.items():
            assumption_rows.append([
                Paragraph(escape(str(key)), cell_small),
                Paragraph(escape(str(item.get("teste", ""))), cell_small),
                _status_cell(item.get("status")),
                _fmt(item.get("statistic")),
                _fmt(item.get("p_value")), Paragraph(escape(str(item.get("mensagem", ""))), caption),
            ])
        story.append(_styled_table(
            assumption_rows,
            col_widths=[2.8 * cm, 2.7 * cm, 2.6 * cm, 1.4 * cm, 1.1 * cm, 5.4 * cm],
        ))
        transform = result.get("transformacao_sugerida")
        if transform:
            story.append(Paragraph(
                f"Transformação sugerida: <b>{escape(str(transform.get('metodo', '—')))}</b> · "
                f"{escape(str(transform.get('descricao', '')))} {escape(str(transform.get('mensagem', '')))}",
                caption,
            ))
        story.append(Spacer(1, 0.35 * cm))

    reg = result.get("regression")
    if reg:
        selected = reg.get("selected_model", {})
        opt = selected.get("optimum") or {}
        reg_text = (
            f"{escape(str(selected.get('equation', '')))} &nbsp;·&nbsp; "
            f"R² ajustado: <b>{_fmt(selected.get('adj_r2'))}</b> &nbsp;·&nbsp; "
            f"AIC: <b>{_fmt(selected.get('aic'))}</b> &nbsp;·&nbsp; "
            f"BIC: <b>{_fmt(selected.get('bic'))}</b> &nbsp;·&nbsp; "
            f"p do termo superior: <b>{_fmt(selected.get('p_top_term'))}</b>"
        )
        if opt.get("x") is not None:
            reg_text += (
                f" &nbsp;·&nbsp; Ponto ótimo estimado: "
                f"<b>x = {_fmt(opt.get('x'))}</b>, y = {_fmt(opt.get('y'))}"
            )
        story.append(KeepTogether([
            section_heading("Regressão"),
            Paragraph(reg_text, body),
        ]))
        lack = selected.get("lack_of_fit") or {}
        story.append(Paragraph(
            (
                f"Falta de ajuste: F={_fmt(lack.get('f_value'))}, p={_fmt(lack.get('p_value'))}. "
                f"{escape(str(lack.get('note', '')))}"
            ),
            caption,
        ))
        reg_png = reg.get("plot_png_base64")
        if reg_png:
            try:
                _rw = 15.0 * cm
                _rh = _rw * 5.2 / 9.0
                _rfig = Image(io.BytesIO(base64.b64decode(reg_png)), width=_rw, height=_rh)
                story.append(KeepTogether([
                    _rfig,
                    Paragraph(
                        "Curva ajustada às médias observadas; linha tracejada indica o ponto "
                        "ótimo estimado.",
                        caption,
                    ),
                ]))
                story.append(Spacer(1, 0.35 * cm))
            except Exception:
                pass

    config = provenance.get("config") or {}
    trace_rows: List[List[Any]] = [["Campo", "Valor"]]
    trace_rows.extend([
        ["Versão do motor", provenance.get("engine_version")],
        ["Commit", provenance.get("git_commit")],
        ["Gerado em Brasília", _format_brasilia_timestamp(provenance.get("generated_at_brasilia") or provenance.get("generated_at_utc"))],
        ["SHA-256 dos dados", provenance.get("data_sha256")],
        ["SHA-256 da configuração", provenance.get("config_sha256")],
    ])
    for key, value in sorted(config.items()):
        trace_rows.append([key, Paragraph(escape(json.dumps(value, ensure_ascii=False, default=str)), caption)])
    story.extend([
        Spacer(1, 0.35 * cm),
        section_heading("Configuração e rastreabilidade"),
        _styled_table(
            trace_rows,
            col_widths=[4.2 * cm, 11.8 * cm],
            right_align_from=None,
            row_padding=5,
        ),
    ])

    doc.build(story)
    return buffer.getvalue()


# ============================================================================
# EXCEL — mantém cabeçalho escuro + corpo claro (impressão amigável)
# ============================================================================
def _style_excel_sheet(worksheet, n_cols: int) -> None:
    header_fill = PatternFill(start_color=HEX["brand_deep"], end_color=HEX["brand_deep"], fill_type="solid")
    header_font = Font(color=HEX["white"], bold=True, size=10, name="Calibri")
    body_font = Font(color=HEX["text_l1"], size=10, name="Calibri")
    zebra_fill = PatternFill(start_color=HEX["surface_subtle"], end_color=HEX["surface_subtle"], fill_type="solid")
    thin = Side(style="thin", color=HEX["surface_line"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, n_cols + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border

    for row in range(2, worksheet.max_row + 1):
        for col in range(1, n_cols + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"
            elif isinstance(cell.value, int):
                cell.number_format = "0"
            if row % 2 == 0:
                cell.fill = zebra_fill

    for col in range(1, n_cols + 1):
        letter = get_column_letter(col)
        max_len = max(
            (len(str(worksheet.cell(row=r, column=col).value or "")) for r in range(1, worksheet.max_row + 1)),
            default=10,
        )
        worksheet.column_dimensions[letter].width = min(max(max_len + 4, 12), 60)

    worksheet.freeze_panes = "A2"
    worksheet.row_dimensions[1].height = 22
    worksheet.sheet_view.showGridLines = False
    if worksheet.max_row > 1 and n_cols > 0:
        worksheet.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{worksheet.max_row}"

    headers = {str(worksheet.cell(1, col).value): col for col in range(1, n_cols + 1)}
    for header in ("significance", "significant", "status"):
        col = headers.get(header)
        if not col:
            continue
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row, col)
            value = str(cell.value or "").lower()
            if value in {"1%", "5%", "true", "violado", "atenção", "atencao"}:
                cell.fill = PatternFill(start_color=HEX["warning_tint"], end_color=HEX["warning_tint"], fill_type="solid")
                cell.font = Font(color=HEX["warning"], bold=True, size=10, name="Calibri")
            elif value in {"false", "ok", "ns"}:
                cell.fill = PatternFill(start_color=HEX["success_tint"], end_color=HEX["success_tint"], fill_type="solid")
                cell.font = Font(color=HEX["success"], bold=True, size=10, name="Calibri")


def build_excel(payload: Dict[str, Any]) -> bytes:
    result = analyze(payload)
    provenance = result.get("provenance") or build_provenance(payload)
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        def write_df(name: str, frame: pd.DataFrame) -> None:
            if frame.empty and len(frame.columns) == 0:
                frame = pd.DataFrame({"informacao": []})
            frame.to_excel(writer, index=False, sheet_name=name)
            _style_excel_sheet(writer.sheets[name], max(1, len(frame.columns)))

        write_df("Resumo", pd.DataFrame({"recomendacao": result.get("recommendations", [])}))

        metadata_rows = [
            {"campo": "Versão do motor", "valor": provenance.get("engine_version")},
            {"campo": "Commit", "valor": provenance.get("git_commit")},
            {"campo": "Gerado em Brasília", "valor": _format_brasilia_timestamp(provenance.get("generated_at_brasilia") or provenance.get("generated_at_utc"))},
            {"campo": "SHA-256 dos dados", "valor": provenance.get("data_sha256")},
            {"campo": "SHA-256 da configuração", "valor": provenance.get("config_sha256")},
        ]
        for key, value in sorted((result.get("meta") or {}).items()):
            metadata_rows.append({"campo": key, "valor": value})
        write_df("Metadados", pd.DataFrame(metadata_rows))

        config_rows = [
            {"campo": key, "valor": json.dumps(value, ensure_ascii=False, default=str)}
            for key, value in sorted((provenance.get("config") or {}).items())
        ]
        write_df("Configuracao", pd.DataFrame(config_rows, columns=["campo", "valor"]))
        write_df("Dados_Entrada", pd.DataFrame(payload.get("data") or []))
        write_df("ANOVA", pd.DataFrame(result.get("anova", {}).get("table", [])))
        means_data = result.get("means", {}).get("treatment_means", [])
        write_df("Medias", pd.DataFrame(means_data))
        means_png_b64 = (result.get("means") or {}).get("plot_png_base64")
        if means_png_b64:
            try:
                _img = XLImage(io.BytesIO(base64.b64decode(means_png_b64)))
                _img.width, _img.height = 640, int(640 * 5.2 / 9)
                writer.sheets["Medias"].add_image(_img, f"A{len(means_data) + 3}")
            except Exception:
                pass

        comparison = (result.get("means") or {}).get("comparison") or {}
        comparison_rows = []
        for row in comparison.get("comparisons") or []:
            comparison_rows.append({
                "test": comparison.get("test"), "alpha": comparison.get("alpha"),
                "control": comparison.get("control"), **row,
            })
        write_df("Comparacoes", pd.DataFrame(comparison_rows, columns=[
            "test", "alpha", "control", "group_a", "group_b", "diff",
            "critical_diff", "p_value", "significant",
        ]))

        factor_rows = []
        for item in result.get("factor_comparisons") or []:
            for level in item.get("levels") or []:
                factor_rows.append({
                    "factor": item.get("factor"), "test": item.get("test"),
                    "alpha": item.get("alpha"), "error_used": item.get("error_used"), **level,
                })
        write_df("Fatores", pd.DataFrame(factor_rows, columns=[
            "factor", "test", "alpha", "error_used", "treatment", "mean", "n", "group",
        ]))

        interaction_rows = []
        for item in result.get("interaction_breakdown") or []:
            for level in item.get("levels") or []:
                interaction_rows.append({
                    "factor": item.get("factor"), "fixed_level": item.get("level"),
                    "sub_factor": item.get("sub_factor"), "test": item.get("test"),
                    "alpha": item.get("alpha"), **level,
                })
        write_df("Interacao", pd.DataFrame(interaction_rows, columns=[
            "factor", "fixed_level", "sub_factor", "test", "alpha", "treatment", "mean", "n", "group",
        ]))
        interaction_png_b64 = (result.get("means") or {}).get("interaction_plot_base64")
        if interaction_png_b64:
            try:
                _iimg = XLImage(io.BytesIO(base64.b64decode(interaction_png_b64)))
                _iimg.width, _iimg.height = 640, int(640 * 5.2 / 9)
                writer.sheets["Interacao"].add_image(_iimg, f"A{len(interaction_rows) + 3}")
            except Exception:
                pass

        assumption_rows = []
        assumption_detail_rows = []
        pressupostos = result.get("pressupostos") or {}
        core_assumption_fields = {"teste", "status", "statistic", "p_value", "n", "mensagem"}
        for key, item in (pressupostos.get("testes") or {}).items():
            assumption_rows.append({
                "pressuposto": key,
                **{field: item.get(field) for field in core_assumption_fields},
            })
            for field, value in item.items():
                if field in core_assumption_fields:
                    continue
                assumption_detail_rows.append({
                    "pressuposto": key,
                    "campo": field,
                    "valor": json.dumps(value, ensure_ascii=False, default=str)
                    if isinstance(value, (dict, list, tuple)) else value,
                })
        write_df("Pressupostos", pd.DataFrame(assumption_rows, columns=[
            "pressuposto", "teste", "status", "statistic", "p_value", "n", "mensagem",
        ]))
        write_df("Pressupostos_Detalhes", pd.DataFrame(assumption_detail_rows, columns=[
            "pressuposto", "campo", "valor",
        ]))

        transform = result.get("transformacao_sugerida")
        write_df("Transformacao", pd.DataFrame([transform] if transform else [], columns=[
            "metodo", "descricao", "aplicado", "mensagem",
        ]))

        regression = result.get("regression") or {}
        regression_rows = []
        for model in regression.get("models") or []:
            lack = model.get("lack_of_fit") or {}
            regression_rows.append({
                "selected": model.get("degree") == regression.get("selected_degree"),
                "degree": model.get("degree"), "equation": model.get("equation"),
                "r2": model.get("r2"), "adj_r2": model.get("adj_r2"),
                "aic": model.get("aic"), "bic": model.get("bic"),
                "p_model": model.get("p_model"), "p_top_term": model.get("p_top_term"),
                "coefficients": json.dumps(model.get("coefficients"), ensure_ascii=False),
                "optimum_x": (model.get("optimum") or {}).get("x"),
                "optimum_y": (model.get("optimum") or {}).get("y"),
                "lack_of_fit_f": lack.get("f_value"), "lack_of_fit_p": lack.get("p_value"),
                "lack_of_fit_significant": lack.get("significant"), "lack_of_fit_note": lack.get("note"),
            })
        write_df("Regressao", pd.DataFrame(regression_rows))

        writer.book.active = 0
    return buffer.getvalue()


# ============================================================================
# GRÁFICOS — matplotlib com tema escuro (idêntico ao site)
# ============================================================================
def _apply_dark_theme(ax) -> None:
    ax.set_facecolor("#0A0A0A")
    ax.figure.patch.set_facecolor("#0A0A0A")
    for spine in ax.spines.values():
        spine.set_color("#262626")
    ax.tick_params(colors="#A3A3A3", which="both")
    ax.xaxis.label.set_color("#A3A3A3")
    ax.yaxis.label.set_color("#A3A3A3")
    ax.title.set_color("#F5F5F5")
    ax.grid(True, color="#1F1F1F", alpha=1.0, linewidth=0.6)


def build_means_plot(payload: Dict[str, Any], fmt: str = "png") -> bytes:
    """Exporta o gráfico de barras das médias por tratamento em PNG ou PDF vetorial, em fundo
    claro (harmonizado com o laudo) e com gradiente de cor por valor: verde nas maiores médias,
    amarelo no meio e laranja nas menores."""
    result = analyze(payload)
    means = result.get("means") or {}
    rows = means.get("treatment_means") or []
    if not rows:
        raise ValueError("Não há médias por tratamento disponíveis para exportar.")

    from matplotlib.colors import LinearSegmentedColormap
    _register_chart_fonts()

    labels = [str(r.get("treatment")) for r in rows]
    vals = [float(r.get("mean") or 0.0) for r in rows]
    sds = [float(r.get("sd") or 0.0) for r in rows]
    groups = [str(r.get("group") or "") for r in rows]
    y_label = str((result.get("meta") or {}).get("response_column") or "Média")

    cmap = LinearSegmentedColormap.from_list(
        "solver_means", ["#E07B39", "#E3B23C", "#7FBF5A", "#2E8B4E"]
    )
    lo, hi = min(vals), max(vals)
    span = (hi - lo) or 1.0
    bar_colors = [cmap((v - lo) / span) for v in vals]

    fig, ax = plt.subplots(figsize=(9, 5.2), dpi=200)
    fig.patch.set_facecolor("#FFFFFF")
    ax.set_facecolor("#FFFFFF")
    ax.set_axisbelow(True)

    x = list(range(len(labels)))
    ax.bar(x, vals, color=bar_colors, edgecolor="#2F2F2F", linewidth=0.6, width=0.66, zorder=2)
    ax.errorbar(x, vals, yerr=sds, fmt="none", ecolor="#5A5A5A", elinewidth=1.3, capsize=4, zorder=3)

    top = max((v + s for v, s in zip(vals, sds)), default=1.0) or 1.0
    for xi, v, s, g in zip(x, vals, sds, groups):
        if g:
            ax.text(xi, v + s + abs(top) * 0.03, g, ha="center", va="bottom",
                    color="#222222", fontsize=11, fontweight="bold")

    ax.set_xticks(x)
    ax.set_xticklabels(labels, color="#333333")
    ax.set_ylabel(y_label, color="#333333")
    ax.set_title("Médias por tratamento", color="#0F5132", pad=14, fontweight="bold", fontname=CHART_TITLE_FONT)
    ax.tick_params(colors="#555555")
    for spine in ax.spines.values():
        spine.set_color("#D6DAD7")
    ax.grid(True, axis="y", color="#E8EBE9")
    ax.margins(y=0.16)

    output = io.BytesIO()
    fig.tight_layout()
    fig.savefig(output, format=fmt, facecolor=fig.get_facecolor(), edgecolor="none")
    plt.close(fig)
    return output.getvalue()


def build_regression_plot(payload: Dict[str, Any], fmt: str = "png") -> bytes:
    """Exporta gráfico de regressão em PNG ou PDF vetorial, em fundo claro (harmonizado com o
    laudo) e com as fontes do documento."""
    result = analyze(payload)
    reg = result.get("regression")
    if not reg:
        raise ValueError("Não há regressão disponível para exportar.")
    _register_chart_fonts()
    points = pd.DataFrame(reg["points"])
    curve = pd.DataFrame(reg["fitted_curve"])
    selected = reg["selected_model"]

    fig, ax = plt.subplots(figsize=(9, 5.2), dpi=200)
    fig.patch.set_facecolor("#FFFFFF")
    ax.set_facecolor("#FFFFFF")
    ax.set_axisbelow(True)

    ax.scatter(points["x"], points["y"], label="Observado", color="#3E9B52",
               edgecolor="#FFFFFF", linewidth=0.7, s=55, zorder=3)
    ax.plot(
        curve["x"], curve["y"],
        label=f"Grau {reg['selected_degree']} · R²aj {selected['adj_r2']:.3f}",
        color="#1F7A3D", linewidth=2.4, zorder=2,
    )
    ax.fill_between(curve["x"], curve["y"], curve["y"].min(), color="#3E9B52", alpha=0.10, zorder=1)

    opt = selected.get("optimum") or {}
    if opt.get("x") is not None:
        ax.axvline(opt["x"], linestyle="--", linewidth=1.1, color="#E07B39", alpha=0.8)
        ax.scatter([opt["x"]], [opt["y"]], marker="o", s=90, label="Ótimo",
                   color="#E07B39", edgecolor="#FFFFFF", linewidth=1.5, zorder=4)

    ax.set_xlabel(reg.get("x_label", "x"), color="#333333")
    ax.set_ylabel(reg.get("y_label", "Resposta"), color="#333333")
    ax.set_title("Regressão ajustada", color="#0F5132", pad=14, fontweight="bold", fontname=CHART_TITLE_FONT)
    ax.tick_params(colors="#555555")
    for spine in ax.spines.values():
        spine.set_color("#D6DAD7")
    ax.grid(True, color="#E8EBE9")
    ax.legend(facecolor="#FFFFFF", edgecolor="#D6DAD7", labelcolor="#333333")

    output = io.BytesIO()
    fig.tight_layout()
    fig.savefig(output, format=fmt, facecolor=fig.get_facecolor(), edgecolor="none")
    plt.close(fig)
    return output.getvalue()
