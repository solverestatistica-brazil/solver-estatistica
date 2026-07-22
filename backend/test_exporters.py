from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook
from pypdf import PdfReader

from exporters import build_excel, build_means_plot, build_pdf
from test_statistics_engine import DBC_EXEMPLO


PAYLOAD = {
    "design": "DBC", "analysis_type": "single", "goal": "max",
    "response_column": "valor", "treatment_column": "tratamento",
    "block_column": "bloco", "comparison_test": "tukey",
    "alpha": 0.05, "alpha_mode": "fixed", "data": DBC_EXEMPLO,
}


def test_pdf_contem_resultados_e_proveniencia():
    pdf = build_pdf(PAYLOAD)
    assert pdf.startswith(b"%PDF")
    assert len(pdf) > 10_000
    reader = PdfReader(io.BytesIO(pdf))
    assert len(reader.pages) >= 2, "o relatório deve ter capa e ao menos uma página de resultados"
    first_page = reader.pages[0]
    width = float(first_page.mediabox.width)
    height = float(first_page.mediabox.height)
    assert height > width
    assert width == pytest.approx(595.28, abs=1.0)
    assert height == pytest.approx(841.89, abs=1.0)

    cover_text = first_page.extract_text() or ""
    assert "RELATÓRIO" in cover_text
    assert "ESTATÍSTICO" in cover_text
    assert "Delineamento Blocos Casualizados (DBC)" in cover_text
    assert "PREPARADO PARA" not in cover_text
    assert "LAUDO TÉCNICO" not in cover_text
    assert "HORÁRIO DE BRASÍLIA" not in cover_text
    assert "Relatório gerado pelo Solver Estatística" not in cover_text
    assert "Fernando Paes Lorena" not in cover_text
    assert "Página" not in cover_text

    body_text = "\n".join(page.extract_text() or "" for page in reader.pages[1:])
    text = cover_text + "\n" + body_text
    assert "Página 1" in body_text
    assert "Gerado em Brasília" in body_text
    assert "BRT" in body_text
    assert "Gerado em UTC" not in body_text
    assert "Alfa adotado" in body_text
    assert "(Fixo)" in body_text
    assert "Delineamento Blocos Casualizados (DBC)" in text
    assert "ATENÇÃO" in body_text
    assert "Heterogeneidade não detectada" in body_text
    assert "diferença de" in body_text
    assert "variâncias" in body_text
    assert "ATENCAO" not in body_text
    for accented_term in ("Relatório", "Brasília", "variação", "observações", "revisão"):
        assert accented_term in text
    assert reader.metadata.title == "Relatório Solver Estatística"
    for expected in (
        "ANOVA", "Comparações detalhadas", "Pressupostos", "Configuração e rastreabilidade",
            "SHA-256 dos dados", "Soma de quadrados",
    ):
        assert expected.upper() in text.upper()


def test_pdf_usa_nome_do_autor_informado():
    payload = {**PAYLOAD, "author_name": "Maria Souza"}
    pdf = build_pdf(payload)
    reader = PdfReader(io.BytesIO(pdf))
    cover_text = reader.pages[0].extract_text() or ""
    assert "Maria Souza" not in cover_text
    assert reader.metadata.author == "Maria Souza"
    assert "Relatório gerado pelo Solver Estatística" not in cover_text


def test_pdf_sem_nome_nao_vaza_nome_pessoal_padrao():
    pdf = build_pdf(PAYLOAD)
    reader = PdfReader(io.BytesIO(pdf))
    cover_text = reader.pages[0].extract_text() or ""
    assert "Fernando Paes Lorena" not in cover_text
    assert reader.metadata.author == "Relatório gerado pelo Solver Estatística"


def test_excel_contem_planilhas_tecnicas_e_dados_de_entrada():
    excel = build_excel(PAYLOAD)
    workbook = load_workbook(io.BytesIO(excel), data_only=False)
    expected = {
        "Resumo", "Metadados", "Configuracao", "Dados_Entrada", "ANOVA", "Medias",
        "Comparacoes", "Pressupostos", "Pressupostos_Detalhes", "Transformacao", "Regressao",
    }
    assert expected.issubset(workbook.sheetnames)
    assert workbook.active.title == "Resumo"
    assert workbook["Dados_Entrada"].max_row == len(DBC_EXEMPLO) + 1
    metadata = {
        workbook["Metadados"].cell(row=row, column=1).value:
        workbook["Metadados"].cell(row=row, column=2).value
        for row in range(2, workbook["Metadados"].max_row + 1)
    }
    assert len(metadata["SHA-256 dos dados"]) == 64
    assert metadata["Versão do motor"]
    assert "Gerado em Brasília" in metadata
    assert "BRT" in metadata["Gerado em Brasília"]
    assert "Gerado em UTC" not in metadata
    assert workbook["ANOVA"].auto_filter.ref
    assert workbook["Resumo"].sheet_view.showGridLines is False


# ---------------------------------------------------------------------------
# Gráfico de médias (ANOVA) — figura de barras com desvio-padrão e letras
# ---------------------------------------------------------------------------
def test_means_plot_embutido_na_analise_e_no_laudo():
    import base64 as _b64
    from statistics_engine import analyze

    result = analyze(PAYLOAD)
    b64 = (result.get("means") or {}).get("plot_png_base64")
    assert b64, "a análise ANOVA deve produzir o gráfico de médias em base64"
    raw = _b64.b64decode(b64)
    assert raw[:8] == b"\x89PNG\r\n\x1a\n", "o gráfico de médias deve ser um PNG válido"
    pdf = build_pdf(PAYLOAD)
    assert pdf.startswith(b"%PDF") and len(pdf) > 10_000


def test_build_means_plot_png_e_pdf_vetorial():
    png = build_means_plot(PAYLOAD, "png")
    assert png[:4] == b"\x89PNG"
    pdf_vec = build_means_plot(PAYLOAD, "pdf")
    assert pdf_vec[:5] == b"%PDF-"


def test_excel_embute_grafico_de_medias():
    import io as _io
    from openpyxl import load_workbook

    xlsx = build_excel(PAYLOAD)
    assert xlsx[:2] == b"PK"
    wb = load_workbook(_io.BytesIO(xlsx))
    assert "Medias" in wb.sheetnames
    assert len(wb["Medias"]._images) == 1, "a aba Medias deve embutir o gráfico de médias"


# ---------------------------------------------------------------------------
# Gráfico de interação (fatorial / split-plot)
# ---------------------------------------------------------------------------
def _fatorial_payload():
    import json as _json
    from pathlib import Path
    data = _json.loads(
        (Path(__file__).resolve().parents[1] / "frontend/assets/data/fatorial_exemplo.json").read_text()
    )
    return {
        "design": "DBC", "analysis_type": "factorial", "goal": "max",
        "response_column": "valor", "block_column": "bloco",
        "factor_columns": ["hibrido", "dose"], "alpha": 0.05, "alpha_mode": "fixed",
        "comparison_test": "tukey", "data": data,
    }


def test_interaction_plot_gerado_e_embutido():
    import base64 as _b64, io as _io
    from openpyxl import load_workbook
    from statistics_engine import analyze

    pay = _fatorial_payload()
    r = analyze(pay)
    ip = (r.get("means") or {}).get("interaction_plot_base64")
    assert ip, "análise fatorial deve gerar o gráfico de interação"
    assert _b64.b64decode(ip)[:8] == b"\x89PNG\r\n\x1a\n"

    wb = load_workbook(_io.BytesIO(build_excel(pay)))
    assert len(wb["Interacao"]._images) == 1, "aba Interacao deve embutir o gráfico"
    assert build_pdf(pay).startswith(b"%PDF")


def test_interaction_plot_ausente_em_fator_unico():
    from statistics_engine import analyze
    r = analyze(PAYLOAD)  # DBC de fator único
    assert "interaction_plot_base64" not in (r.get("means") or {})
